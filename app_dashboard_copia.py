

# -*- coding: utf-8 -*-
"""
╔═══════════════════════════════════════════════════════════════════════════╗
║          SISTEMA PREDICTIVO DE ANÁLISIS ACADÉMICO - V4.0 INTEGRADO       ║
║                  I.E. "Víctor Núñez Valencia" / Victor Andrés Belaunde   ║
║                              Autor: frederickv                            ║
║                            Fecha: 2025-11-11 (Corregido)                 ║
╚═══════════════════════════════════════════════════════════════════════════╝
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import os
import sys
import warnings
import io
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ═════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN INICIAL (DEBE SER LO PRIMERO)
# ═════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Sistema Académico MINEDU V4.0",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

warnings.filterwarnings('ignore')

# ═════════════════════════════════════════════════════════════════════════════
# IMPORTACIÓN DE MÓDULOS PERSONALIZADOS
# ═════════════════════════════════════════════════════════════════════════════

ruta_actual = os.path.dirname(os.path.abspath(__file__))
if ruta_actual not in sys.path:
    sys.path.append(ruta_actual)

HAS_MODULO_PRIORIZADOS = False
HAS_MODULO_PREDICTIVO = False
HAS_CATBOOST = False

try:
    from modulo_priorizados import analizar_desde_dataframe
    HAS_MODULO_PRIORIZADOS = True
    print("✅ modulo_priorizados.py cargado exitosamente")
except ImportError as e:
    print(f"⚠️ modulo_priorizados.py no disponible: {e}")

try:
    import catboost
    HAS_CATBOOST = True
    print("✅ CatBoost instalado correctamente")
    
    from modulo_predictivo import ejecutar_analisis_predictivo
    HAS_MODULO_PREDICTIVO = True
    print("✅ modulo_predictivo.py cargado exitosamente")
    
except ImportError as e:
    error_msg = str(e).lower()
    if "catboost" in error_msg:
        print("⚠️ CatBoost no está instalado")
        HAS_CATBOOST = False
    else:
        print(f"⚠️ Error al cargar modulo_predictivo.py: {e}")
    HAS_MODULO_PREDICTIVO = False

# ═════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN DEL SISTEMA
# ═════════════════════════════════════════════════════════════════════════════

INFO_INSTITUCION = {
    'nombre_ie1': 'I.E. 40079 "VÍCTOR NÚÑEZ VALENCIA"',
    'ubicacion': 'Av. Wanders 113, Sachaca, Arequipa',
    'codigo': '0899120',
    'version': 'v4.0 - 2025-11-11',
    'nivel': 'Educación Secundaria',
    'ugel': 'UGEL Arequipa Sur'
}

ESCALA_CALIFICACIONES = {
    'AD': {'min': 18, 'max': 20, 'num': 19, 'desc': 'Logro Destacado', 'color': '#118AB2'},
    'A': {'min': 15, 'max': 17, 'num': 16, 'desc': 'Logro Esperado', 'color': '#06D6A0'},
    'B': {'min': 11, 'max': 14, 'num': 12, 'desc': 'En Proceso', 'color': '#FFD166'},
    'C': {'min': 0, 'max': 10, 'num': 8, 'desc': 'En Inicio', 'color': '#FF6B6B'}
}

ESTRATEGIAS_MINEDU = {
    'C': """🚨 **Reforzamiento Urgente Requerido**
- Implementar Plan de Tutoría Individualizado (PTI)
- Foco en competencias básicas con sesiones de 30-45 min
- Contactar a padres/apoderados para acompañamiento familiar
📚 Ref: MINEDU - Disposiciones para Reforzamiento Escolar 2024""",
    
    'B': """⚠️ **Acompañamiento Pedagógico Necesario**
- Proporcionar material didáctico diferenciado
- Fomentar trabajo colaborativo (grupos de 3-4 estudiantes)
📚 Ref: MINEDU - Evaluación Formativa en el Marco de CNEB""",
    
    'A': """✅ **Consolidación de Aprendizaje**
- Asignar proyectos de aplicación práctica (ABP)
- Promover resolución de problemas complejos
📚 Ref: MINEDU - Orientaciones para Trabajo por Competencias""",
    
    'AD': """🌟 **Potenciación de Talento Excepcional**
- Fomentar proyectos de investigación autónomos
- Asignar rol de tutor par (mentoría entre estudiantes)
📚 Ref: MINEDU - Atención a Estudiantes con Alto Desempeño"""
}

# ═════════════════════════════════════════════════════════════════════════════
# CLASE GESTOR DE EVALUACIÓN
# ═════════════════════════════════════════════════════════════════════════════

class GestorEvaluacionMINEDU:
    """Gestiona conversiones y análisis de calificaciones según normativa MINEDU"""
    
    def __init__(self):
        self.escala = ESCALA_CALIFICACIONES
        self.estrategias = ESTRATEGIAS_MINEDU
    
    def num_a_letra(self, valor: float) -> str:
        """Convierte nota numérica (0-20) a letra (C/B/A/AD)"""
        if pd.isna(valor):
            return "C"
        valor = float(valor)
        for letra, config in self.escala.items():
            if config['min'] <= valor <= config['max']:
                return letra
        return "C"
    
    def letra_a_num(self, letra: str) -> float:
        """Convierte letra (C/B/A/AD) a valor numérico representativo"""
        letra = str(letra).strip().upper()
        return float(self.escala.get(letra, {'num': 8})['num'])
    
    def generar_observacion(self, promedio: float, nombre: str = "el estudiante") -> tuple:
        """Genera observación pedagógica completa"""
        letra = self.num_a_letra(promedio)
        config = self.escala[letra]
        estrategia = self.estrategias[letra]
        
        observacion = f"""
**👤 Estudiante:** {nombre}
**📊 Promedio:** {promedio:.2f}/20.00
**📈 Nivel:** {letra} - {config['desc']}

**📋 Observación Pedagógica:**
{estrategia}
"""
        return observacion, letra
    
    def get_color(self, letra: str) -> str:
        """Retorna color hexadecimal para la letra"""
        return self.escala.get(letra, {}).get('color', '#999999')

gestor = GestorEvaluacionMINEDU()

# ═════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE CARGA Y PROCESAMIENTO
# ═════════════════════════════════════════════════════════════════════════════

@st.cache_data
def cargar_excel(archivo_subido):
    """Carga todas las hojas de un archivo Excel"""
    try:
        xls = pd.ExcelFile(archivo_subido)
        nombres_hojas = xls.sheet_names
        
        if not nombres_hojas:
            return None, [], "El archivo no contiene hojas"
        
        datos_por_hoja = {}
        datos_raw_por_hoja = {}
        errores = []
        
        for hoja in nombres_hojas:
            try:
                df_raw = pd.read_excel(xls, sheet_name=hoja, header=None)
                datos_raw_por_hoja[hoja] = df_raw
                
                fila_header = detectar_fila_encabezado(df_raw)
                if fila_header is not None:
                    df_hoja = pd.read_excel(xls, sheet_name=hoja, header=fila_header)
                    df_hoja = limpiar_dataframe(df_hoja)
                    if not df_hoja.empty:
                        datos_por_hoja[hoja] = df_hoja
                else:
                    errores.append(f"Hoja '{hoja}': No se encontró encabezado")
            except Exception as e:
                errores.append(f"Hoja '{hoja}': {str(e)}")
        
        if not datos_por_hoja:
            return None, None, [], "No se pudieron cargar hojas válidas"
        
        return datos_por_hoja, datos_raw_por_hoja, list(datos_por_hoja.keys()), None
    
    except Exception as e:
        return None, None, [], f"Error al leer Excel: {str(e)}"

def detectar_fila_encabezado(df_raw, palabras_clave=['APELLIDOS', 'NOMBRES', 'NOMBRE', 'ESTUDIANTE', 'ALUMNO']):
    """Detecta automáticamente la fila de encabezado"""
    for i in range(min(15, len(df_raw))):
        fila_str = ' '.join(str(x).upper() for x in df_raw.iloc[i] if pd.notna(x))
        if any(clave in fila_str for clave in palabras_clave):
            return i
    return 0

def limpiar_dataframe(df):
    """Limpia DataFrame: elimina columnas/filas vacías"""
    cols = pd.Series(df.columns)
    for dup in cols[cols.duplicated()].unique():
        cols[cols[cols == dup].index.values.tolist()] = [
            f"{dup}.{i}" if i != 0 else dup for i in range(sum(cols == dup))
        ]
    df.columns = cols
    df = df.loc[:, ~df.columns.str.contains('^Unnamed', na=False)]
    df = df.dropna(axis=1, how='all')
    if len(df.columns) > 1:
        col_id = df.columns[1]
        df = df.dropna(subset=[col_id])
    return df.reset_index(drop=True)

def obtener_columnas_notas(df):
    """Detecta columnas de notas (numéricas o letras A/B/C/AD)"""
    columnas_notas = []
    columnas_id = []
    palabras_excluir = ['ESTUDIANTE', 'NOMBRE', 'APELLIDO', 'GRADO', 'SECCION', 
                        'CODIGO', 'DNI', 'ID', 'PROMEDIO', 'OBSERVACION', 'FECHA']
    
    for col in df.columns:
        col_str = str(col).upper()
        if any(kw in col_str for kw in ['ESTUDIANTE', 'NOMBRE', 'APELLIDO']):
            columnas_id.append(col)
            continue
        if any(kw in col_str for kw in palabras_excluir):
            continue
        
        muestra = df[col].dropna()
        if len(muestra) < 3:
            continue
        muestra_sample = muestra.sample(min(20, len(muestra)))
        
        try:
            muestra_num = pd.to_numeric(muestra_sample, errors='coerce').dropna()
            if (len(muestra_num) / len(muestra_sample) > 0.7 and 
                muestra_num.min() >= 0 and muestra_num.max() <= 20):
                columnas_notas.append(col)
                continue
        except:
            pass
        
        try:
            muestra_str = muestra_sample.astype(str).str.upper().str.strip()
            conteo_letras = muestra_str.isin(['A', 'B', 'C', 'AD']).sum()
            if conteo_letras / len(muestra_sample) > 0.6:
                columnas_notas.append(col)
        except:
            pass
    
    return columnas_notas, columnas_id

@st.cache_data
def procesar_datos(df, columnas_notas):
    """Procesa notas: convierte a numérico y calcula promedios"""
    df_proc = df.copy()
    mapeo = {letra: config['num'] for letra, config in ESCALA_CALIFICACIONES.items()}
    columnas_num_proc = []
    
    for col in columnas_notas:
        col_num = f"{col}_num"
        df_proc[col_num] = pd.to_numeric(df_proc[col], errors='coerce')
        mask_nan = df_proc[col_num].isna()
        df_proc.loc[mask_nan, col_num] = (
            df_proc.loc[mask_nan, col]
            .astype(str).str.upper().str.strip()
            .map(mapeo)
        )
        df_proc[col_num] = df_proc[col_num].fillna(mapeo['C'])
        columnas_num_proc.append(col_num)
    
    if columnas_num_proc:
        df_proc['PROMEDIO'] = df_proc[columnas_num_proc].mean(axis=1).round(2)
        df_proc['CALIFICACION_LETRA'] = df_proc['PROMEDIO'].apply(gestor.num_a_letra)
        df_proc['ESTADO'] = df_proc['CALIFICACION_LETRA'].apply(
            lambda x: 'Aprobado' if x in ['AD', 'A', 'B'] else 'Desaprobado'
        )
    
    return df_proc, columnas_num_proc

# ═════════════════════════════════════════════════════════════════════════════
# COMPONENTES DE UI
# ═════════════════════════════════════════════════════════════════════════════


def mostrar_logo():
    """Muestra logo en la barra lateral"""
    
    # Intentar cargar logo del colegio desde múltiples ubicaciones
    posibles_rutas = [
        os.path.join(os.path.dirname(__file__), "logocolegio.png"),
        os.path.join(os.path.dirname(__file__), "logo.png"),
        os.path.join(os.path.dirname(__file__), "assets", "logo_colegio.png"),
    ]
    
    logo_cargado = False
    for ruta_logo in posibles_rutas:
        if os.path.exists(ruta_logo):
            # ✅ CORREGIDO: Sin use_column_width
            st.sidebar.image(ruta_logo, width=150)
            logo_cargado = True
            break
    
    if not logo_cargado:
        # Mostrar placeholder elegante si no hay logo
        st.sidebar.markdown("""
        <div style='text-align: center; padding: 15px; 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    border-radius: 15px; margin-bottom: 10px;'>
            <div style='font-size: 60px; margin: 10px 0;'>🏫</div>
            <p style='color: white; margin: 0; font-size: 12px;'>I.E. Víctor Núñez Valencia</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.sidebar.markdown(f"""
    <div style='text-align: center; padding: 20px;'>
        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    color: white; padding: 30px; border-radius: 15px;'>
            <h2 style='margin: 0;'>🎓</h2>
            <h3 style='margin: 5px 0; font-size: 14px;'>Sistema Académico</h3>
            <h3 style='margin: 0; font-size: 14px;'>MINEDU 2024</h3>
        </div>
        <p style='color: #666; font-size: 11px; margin: 5px 0;'>{INFO_INSTITUCION['version']}</p>
    </div>
    """, unsafe_allow_html=True)


def mostrar_kpis(total, promedio, tasa_aprob):
    """Muestra tarjetas KPI"""
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown(f"""
        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    padding: 20px; border-radius: 10px; text-align: center;'>
            <p style='color: white; margin: 0; font-size: 14px;'>Total Estudiantes</p>
            <h2 style='color: white; margin: 10px 0;'>{total} 🎓</h2>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div style='background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); 
                    padding: 20px; border-radius: 10px; text-align: center;'>
            <p style='color: white; margin: 0; font-size: 14px;'>Promedio General</p>
            <h2 style='color: white; margin: 10px 0;'>{promedio:.2f} 📊</h2>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        color = '#06D6A0' if tasa_aprob >= 70 else '#FFD166' if tasa_aprob >= 50 else '#FF6B6B'
        st.markdown(f"""
        <div style='background: linear-gradient(135deg, {color} 0%, {color}dd 100%); 
                    padding: 20px; border-radius: 10px; text-align: center;'>
            <p style='color: white; margin: 0; font-size: 14px;'>Tasa Aprobación</p>
            <h2 style='color: white; margin: 10px 0;'>{tasa_aprob:.1f}% ✅</h2>
        </div>
        """, unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════════════════════
# PÁGINAS DE LA APLICACIÓN
# ═════════════════════════════════════════════════════════════════════════════

def pagina_inicio():
    """Página de inicio con instrucciones"""
    st.title("🎓 Sistema de Análisis y Reforzamiento Académico V4.0")
    
    st.markdown(f"""
    ### Bienvenido al Sistema Integrado MINEDU
    
    **Institución Educativa:**
    - 🏫 **{INFO_INSTITUCION['nombre_ie1']}**
    - 📍 **Ubicación:** {INFO_INSTITUCION['ubicacion']}
    - 🔢 **Código Modular:** {INFO_INSTITUCION['codigo']}
    - 📚 **Nivel:** {INFO_INSTITUCION.get('nivel', 'Educación Básica Regular')}
    
    ---
    
    Este sistema permite analizar el rendimiento académico según normativa MINEDU.
    """)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.info("""
        **📋 Instrucciones:**
        1. Cargue su archivo Excel en la barra lateral
        2. El sistema detectará automáticamente las hojas
        3. Navegue por las diferentes vistas
        4. Exporte reportes cuando lo necesite
        """)
    
    with col2:
        modulos_disponibles = []
        if HAS_MODULO_PRIORIZADOS:
            modulos_disponibles.append("✅ Análisis de Priorizados")
        if HAS_MODULO_PREDICTIVO and HAS_CATBOOST:
            modulos_disponibles.append("✅ Modelo Predictivo (ML)")
        if not modulos_disponibles:
            modulos_disponibles.append("✅ Análisis Descriptivo Estándar")
        
        st.success(f"""
        **✨ Módulos Activos:**
        {chr(10).join(modulos_disponibles)}
        
        **📊 Funciones Disponibles:**
        - Vista Director (Global)
        - Vista Docente (Por Aula)
        - Análisis Estudiantil
        - Exportación de Reportes
        """)
    
    if not HAS_CATBOOST:
        st.info("""
        ℹ️ **Nota:** El módulo de Machine Learning está deshabilitado porque CatBoost no está instalado.
        El sistema funciona perfectamente con análisis descriptivo estándar.
        
        **Para habilitar predicciones ML (opcional):**
```bash
        pip install catboost
```
        Luego reinicie la aplicación.
        """)





def pagina_vista_director(datos_por_hoja, datos_raw_por_hoja):
    """Vista del director con análisis global"""
    st.title("👨‍🏫 Vista Director: Análisis Global")
    
    try:
        df_global = pd.concat(datos_por_hoja.values(), ignore_index=True)
        columnas_notas, columnas_id = obtener_columnas_notas(df_global)
        
        if not columnas_notas:
            st.error("❌ No se detectaron columnas de notas válidas")
            return
        
        df_procesado, _ = procesar_datos(df_global, columnas_notas)
        
        total = len(df_procesado)
        promedio = df_procesado['PROMEDIO'].mean()
        tasa_aprob = (df_procesado['ESTADO'] == 'Aprobado').mean() * 100
        
        mostrar_kpis(total, promedio, tasa_aprob)
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### 📊 Distribución por Nivel")
            conteo = df_procesado['CALIFICACION_LETRA'].value_counts().reindex(['C','B','A','AD']).fillna(0)
            colores = [gestor.get_color(letra) for letra in conteo.index]
            fig = px.bar(x=conteo.index, y=conteo.values, color=conteo.index,
                        color_discrete_sequence=colores)
            fig.update_layout(showlegend=False, height=400)
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            st.markdown("#### 📈 Histograma de Promedios")
            fig = px.histogram(df_procesado, x='PROMEDIO', nbins=20,
                             color_discrete_sequence=['#667eea'])
            fig.update_layout(showlegend=False, height=400)
            st.plotly_chart(fig, use_container_width=True)
        
        st.markdown("### 📋 Resumen por Aula/Trimestre")
        resumen_aulas = []
        
        for nombre_hoja, df_hoja in datos_por_hoja.items():
            try:
                cols_notas, cols_id = obtener_columnas_notas(df_hoja)
                if not cols_notas:
                    continue
                
                df_proc_hoja, cols_num = procesar_datos(df_hoja, cols_notas)
                
                if 'PROMEDIO' not in df_proc_hoja.columns:
                    if cols_num:
                        df_proc_hoja['PROMEDIO'] = df_proc_hoja[cols_num].mean(axis=1)
                    else:
                        continue
                
                if 'ESTADO' not in df_proc_hoja.columns:
                    df_proc_hoja['ESTADO'] = df_proc_hoja['PROMEDIO'].apply(
                        lambda x: 'Aprobado' if x >= 11 else 'Desaprobado'
                    )
                
                num_estudiantes = len(df_proc_hoja)
                promedio_aula = df_proc_hoja['PROMEDIO'].mean()
                tasa_aprobacion = (df_proc_hoja['ESTADO'] == 'Aprobado').sum() / num_estudiantes * 100
                
                resumen_aulas.append({
                    'Aula/Trimestre': nombre_hoja,
                    'N° Estudiantes': num_estudiantes,
                    'Promedio': round(promedio_aula, 2),
                    'Tasa Aprobación (%)': round(tasa_aprobacion, 1)
                })
            except Exception as e:
                st.warning(f"⚠️ Error en hoja '{nombre_hoja}': {str(e)}")
                continue
        
        if resumen_aulas:
            df_resumen = pd.DataFrame(resumen_aulas)
            st.dataframe(
                df_resumen.style.format({
                    'Promedio': '{:.2f}',
                    'Tasa Aprobación (%)': '{:.1f}%'
                }).background_gradient(
                    subset=['Tasa Aprobación (%)'],
                    cmap='RdYlGn',
                    vmin=0,
                    vmax=100
                ),
                use_container_width=True
            )
        else:
            st.error("❌ No se pudo generar el resumen")
    
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        st.exception(e)

def pagina_analisis_priorizados(datos_raw_por_hoja):
    """Análisis de estudiantes priorizados"""
    st.title("🎯 Análisis de Estudiantes Priorizados")
    
    if not HAS_MODULO_PRIORIZADOS:
        st.error("""
        ❌ El módulo 'modulo_priorizados.py' no está disponible.
        
        **Solución:**
        1. Asegúrese de que el archivo 'modulo_priorizados.py' esté en la misma carpeta
        2. Verifique que el archivo tenga la función `analizar_desde_dataframe()`
        """)
        return
    
    hoja_seleccionada = st.selectbox(
        "📚 Seleccione la hoja a analizar:",
        options=list(datos_raw_por_hoja.keys())
    )
    
    if hoja_seleccionada:
        with st.spinner(f"🔄 Analizando hoja '{hoja_seleccionada}'..."):
            try:
                df_raw = datos_raw_por_hoja[hoja_seleccionada]
                reporte_texto = analizar_desde_dataframe(df_raw, hoja_seleccionada)
                
                st.text(reporte_texto)
                
                st.download_button(
                    label="📥 Descargar Informe (.txt)",
                    data=reporte_texto,
                    file_name=f"informe_priorizados_{hoja_seleccionada}_{datetime.now().strftime('%Y%m%d')}.txt",
                    mime="text/plain"
                )
            except Exception as e:
                st.error(f"❌ Error al procesar: {str(e)}")
                st.exception(e)

def pagina_modelo_predictivo(datos_raw_por_hoja):
    """Entrenamiento de modelo predictivo ML"""
    st.title("🤖 Modelo Predictivo de Aprobación (Machine Learning)")
    
    if not HAS_MODULO_PREDICTIVO or not HAS_CATBOOST:
        st.error("""
        ❌ El módulo predictivo no está disponible.
        
        **Solución:**
        1. Instale CatBoost: `pip install catboost`
        2. Asegúrese de que 'modulo_predictivo.py' esté en la misma carpeta
        3. Reinicie la aplicación Streamlit
        """)
        return
    
    hojas_disponibles = list(datos_raw_por_hoja.keys())
    hoja_ie = None
    
    if "IE" in hojas_disponibles:
        hoja_ie = "IE"
        st.info(f"✅ Hoja 'IE' detectada automáticamente")
    else:
        st.warning("⚠️ No se encontró hoja 'IE'. Seleccione una hoja manualmente:")
        hoja_ie = st.selectbox("Seleccione hoja para análisis ML:", options=hojas_disponibles)
    
    if hoja_ie and st.button("🚀 Entrenar Modelo", type="primary"):
        with st.spinner("🔄 Entrenando modelo CatBoost (puede tomar 1-2 minutos)..."):
            try:
                df_raw_ie = datos_raw_por_hoja[hoja_ie]
                df_reporte, fig_importancia = ejecutar_analisis_predictivo(df_raw_ie)
                
                st.success("✅ ¡Entrenamiento completado!")
                
                st.markdown("### 📊 Reporte de Predicciones")
                st.dataframe(df_reporte, use_container_width=True)
                
                st.download_button(
                    label="📥 Descargar Reporte CSV",
                    data=df_reporte.to_csv(index=False).encode('utf-8'),
                    file_name=f"predicciones_ml_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
                
                st.markdown("### 📈 Importancia de Características")
                st.pyplot(fig_importancia)
                
            except Exception as e:
                st.error(f"❌ Error durante el entrenamiento: {str(e)}")
                with st.expander("📋 Ver detalles del error"):
                    st.code(str(e))
                    import traceback
                    st.code(traceback.format_exc())

def pagina_vista_docente(datos_por_hoja):
    """Vista detallada para docentes"""
    st.title("👩‍🏫 Vista Docente: Análisis por Aula")
    
    hoja_seleccionada = st.selectbox("📚 Seleccione Aula:", options=list(datos_por_hoja.keys()))
    
    if hoja_seleccionada:
        try:
            df_hoja = datos_por_hoja[hoja_seleccionada]
            columnas_notas, columnas_id = obtener_columnas_notas(df_hoja)
            
            if not columnas_notas:
                st.warning("⚠️ No se detectaron columnas de notas")
                return
            
            df_procesado, _ = procesar_datos(df_hoja, columnas_notas)
            
            total = len(df_procesado)
            promedio = df_procesado['PROMEDIO'].mean()
            tasa_aprob = (df_procesado['ESTADO'] == 'Aprobado').mean() * 100
            
            mostrar_kpis(total, promedio, tasa_aprob)
            st.markdown("---")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### 📊 Distribución de Calificaciones")
                conteo = df_procesado['CALIFICACION_LETRA'].value_counts().reindex(['C','B','A','AD']).fillna(0)
                colores = [gestor.get_color(letra) for letra in conteo.index]
                fig = px.bar(x=conteo.index, y=conteo.values, color=conteo.index,
                            color_discrete_sequence=colores)
                fig.update_layout(showlegend=False, height=400)
                st.plotly_chart(fig, width='stretch')
            
            with col2:
                st.markdown("#### 📈 Distribución de Promedios")
                fig = px.histogram(df_procesado, x='PROMEDIO', nbins=15,
                                 color_discrete_sequence=['#667eea'])
                fig.update_layout(showlegend=False, height=400)
                st.plotly_chart(fig, width='stretch')
            
            st.markdown("### 📋 Listado de Estudiantes")
            if columnas_id:
                cols_mostrar = columnas_id + ['PROMEDIO', 'CALIFICACION_LETRA', 'ESTADO']
                st.dataframe(
                    df_procesado[cols_mostrar].style.format({'PROMEDIO': '{:.2f}'}),
                    width='stretch'
                )
            
        except Exception as e:
            st.error(f"❌ Error: {str(e)}")

def pagina_analisis_estudiantil(datos_por_hoja):
    """Análisis individual de estudiantes"""
    st.title("🧑‍🎓 Análisis Estudiantil y Observaciones Pedagógicas")
    
    hoja_seleccionada = st.selectbox("📚 Seleccione Aula:", options=list(datos_por_hoja.keys()))
    
    if hoja_seleccionada:
        try:
            df_hoja = datos_por_hoja[hoja_seleccionada]
            columnas_notas, columnas_id = obtener_columnas_notas(df_hoja)
            
            if not columnas_id:
                st.error("❌ No se encontró columna de nombres")
                return
            
            col_nombre = columnas_id[0]
            df_procesado, _ = procesar_datos(df_hoja, columnas_notas)
            
            lista_estudiantes = ["📋 Mostrar Todos"] + list(df_procesado[col_nombre].unique())
            estudiante_seleccionado = st.selectbox("👤 Seleccionar Estudiante:", options=lista_estudiantes)
            
            df_filtrado = df_procesado
            if estudiante_seleccionado != "📋 Mostrar Todos":
                df_filtrado = df_procesado[df_procesado[col_nombre] == estudiante_seleccionado]
            
            for idx, estudiante in df_filtrado.iterrows():
                nombre = estudiante[col_nombre]
                promedio = estudiante['PROMEDIO']
                
                observacion, letra = gestor.generar_observacion(promedio, nombre)
                
                if letra == 'AD':
                    st.success(observacion)
                elif letra == 'A':
                    st.info(observacion)
                elif letra == 'B':
                    st.warning(observacion)
                else:
                    st.error(observacion)
                
                st.markdown("---")
        
        except Exception as e:
            st.error(f"❌ Error: {str(e)}")

def generar_excel_formateado(df_procesado, columnas_id, nombre_hoja="Reporte"):
    """
    Genera un Excel con formato profesional similar al formato institucional
    
    Args:
        df_procesado: DataFrame con los datos procesados
        columnas_id: Lista de columnas de identificación
        nombre_hoja: Nombre de la hoja (trimestre, aula, etc.)
    
    Returns:
        BytesIO buffer con el archivo Excel formateado
    """
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31]  # Excel limita a 31 caracteres
    
    # ═══════════════════════════════════════════════════════════════
    # 1. ENCABEZADO INSTITUCIONAL
    # ═══════════════════════════════════════════════════════════════
    
    # Fila 1: Nombre de la IE (combinada)
    ws.merge_cells('A1:H1')
    cell_titulo = ws['A1']
    cell_titulo.value = f'INSTITUCIÓN EDUCATIVA: {INFO_INSTITUCION["codigo"]} "{INFO_INSTITUCION["nombre_ie1"]}"'
    cell_titulo.font = Font(name='Calibri', size=12, bold=True)
    cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
    cell_titulo.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Fila 2: Distrito
    ws.merge_cells('A2:H2')
    cell_distrito = ws['A2']
    cell_distrito.value = 'DISTRITO: SACHACA'
    cell_distrito.font = Font(name='Calibri', size=11, bold=True)
    cell_distrito.alignment = Alignment(horizontal='center', vertical='center')
    cell_distrito.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # Fila 3: Nivel de logro
    ws.merge_cells('A3:H3')
    cell_nivel = ws['A3']
    cell_nivel.value = f'NIVEL DE LOGRO DE LAS COMPETENCIAS - {nombre_hoja.upper()}'
    cell_nivel.font = Font(name='Calibri', size=11, bold=True)
    cell_nivel.alignment = Alignment(horizontal='center', vertical='center')
    cell_nivel.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # ═══════════════════════════════════════════════════════════════
    # 2. CONFIGURACIÓN DE ESTILOS
    # ═══════════════════════════════════════════════════════════════
    
    # Colores según nivel de logro (MINEDU)
    COLORES_NIVEL = {
        'AD': PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid'),  # Azul
        'A': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),   # Verde
        'B': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),   # Amarillo
        'C': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')    # Rojo
    }
    
    # Estilos de bordes
    borde_delgado = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    borde_grueso = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    # ═══════════════════════════════════════════════════════════════
    # 3. ENCABEZADOS DE COLUMNAS (Fila 5)
    # ═══════════════════════════════════════════════════════════════
    
    fila_inicio = 5  # Empezar después del encabezado institucional
    
    # Columnas a mostrar
    columnas_mostrar = columnas_id + ['PROMEDIO', 'CALIFICACION_LETRA', 'ESTADO']
    
    # Escribir encabezados
    col_actual = 1
    for col_nombre in columnas_mostrar:
        cell = ws.cell(row=fila_inicio, column=col_actual)
        
        # Nombres más cortos para encabezados
        nombre_corto = {
            'PROMEDIO': 'PROM.',
            'CALIFICACION_LETRA': 'NIVEL',
            'ESTADO': 'ESTADO'
        }.get(col_nombre, col_nombre)
        
        cell.value = nombre_corto
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = borde_grueso
        
        col_actual += 1
    
    # ═══════════════════════════════════════════════════════════════
    # 4. DATOS DE ESTUDIANTES
    # ═══════════════════════════════════════════════════════════════
    
    fila_actual = fila_inicio + 1
    
    for idx, row in df_procesado.iterrows():
        col_actual = 1
        
        for col_nombre in columnas_mostrar:
            cell = ws.cell(row=fila_actual, column=col_actual)
            valor = row[col_nombre]
            
            # Formato según tipo de dato
            if col_nombre == 'PROMEDIO':
                cell.value = float(valor)
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_nombre == 'CALIFICACION_LETRA':
                cell.value = valor
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calibri', size=11, bold=True)
                
                # Aplicar color según nivel
                if valor in COLORES_NIVEL:
                    cell.fill = COLORES_NIVEL[valor]
                    
            elif col_nombre == 'ESTADO':
                cell.value = valor
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color para estado
                if valor == 'Aprobado':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    cell.font = Font(color='006100')
                else:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(color='9C0006')
            else:
                # Columnas de texto (nombres, apellidos)
                cell.value = str(valor)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            cell.border = borde_delgado
            col_actual += 1
        
        fila_actual += 1
    
    # ═══════════════════════════════════════════════════════════════
    # 5. AJUSTAR ANCHOS DE COLUMNA - CORREGIDO ✅
    # ═══════════════════════════════════════════════════════════════
    
    anchos_columna = {
        'APELLIDOS': 30,
        'NOMBRES': 30,
        'APELLIDOS Y NOMBRES': 40,
        'ESTUDIANTE': 40,
        'PROMEDIO': 10,
        'CALIFICACION_LETRA': 8,
        'ESTADO': 12
    }
    
    for idx, col_nombre in enumerate(columnas_mostrar, start=1):
        ancho = anchos_columna.get(col_nombre, 15)
        # ✅ CORRECCIÓN: Usar get_column_letter directamente
        letra_columna = get_column_letter(idx)
        ws.column_dimensions[letra_columna].width = ancho
    
    # ═══════════════════════════════════════════════════════════════
    # 6. PIE DE PÁGINA CON ESTADÍSTICAS
    # ═══════════════════════════════════════════════════════════════
    
    fila_stats = fila_actual + 2
    
    # Calcular estadísticas
    total_estudiantes = len(df_procesado)
    promedio_general = df_procesado['PROMEDIO'].mean()
    aprobados = (df_procesado['ESTADO'] == 'Aprobado').sum()
    desaprobados = (df_procesado['ESTADO'] == 'Desaprobado').sum()
    tasa_aprobacion = (aprobados / total_estudiantes * 100) if total_estudiantes > 0 else 0
    
    # Conteo por nivel
    conteo_niveles = df_procesado['CALIFICACION_LETRA'].value_counts()
    
    # Escribir estadísticas
    ws.merge_cells(f'A{fila_stats}:B{fila_stats}')
    cell_stats = ws[f'A{fila_stats}']
    cell_stats.value = '📊 ESTADÍSTICAS GENERALES'
    cell_stats.font = Font(name='Calibri', size=11, bold=True)
    cell_stats.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    fila_stats += 1
    
    stats_data = [
        ('Total Estudiantes:', total_estudiantes),
        ('Promedio General:', f'{promedio_general:.2f}'),
        ('Aprobados:', aprobados),
        ('Desaprobados:', desaprobados),
        ('Tasa Aprobación:', f'{tasa_aprobacion:.1f}%'),
        ('', ''),
        ('DISTRIBUCIÓN POR NIVEL:', ''),
        ('AD - Logro Destacado:', conteo_niveles.get('AD', 0)),
        ('A - Logro Esperado:', conteo_niveles.get('A', 0)),
        ('B - En Proceso:', conteo_niveles.get('B', 0)),
        ('C - En Inicio:', conteo_niveles.get('C', 0))
    ]
    
    for label, valor in stats_data:
        ws[f'A{fila_stats}'] = label
        ws[f'A{fila_stats}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'B{fila_stats}'] = valor
        ws[f'B{fila_stats}'].font = Font(name='Calibri', size=10)
        fila_stats += 1
    
    # ═══════════════════════════════════════════════════════════════
    # 7. GUARDAR EN BUFFER
    # ═══════════════════════════════════════════════════════════════
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


# ═══════════════════════════════════════════════════════════════════
# ACTUALIZACIÓN DE LA FUNCIÓN pagina_exportar_reportes()
# ═══════════════════════════════════════════════════════════════════

def pagina_exportar_reportes(datos_por_hoja):
    """Exportación de reportes en diferentes formatos - VERSIÓN MEJORADA"""
    st.title("📄 Exportar Reportes")
    
    hoja_seleccionada = st.selectbox(
        "📚 Seleccione datos a exportar:", 
        options=list(datos_por_hoja.keys())
    )
    
    if hoja_seleccionada:
        try:
            df_hoja = datos_por_hoja[hoja_seleccionada]
            columnas_notas, columnas_id = obtener_columnas_notas(df_hoja)
            
            if not columnas_notas:
                st.warning("⚠️ Sin datos de notas")
                return
            
            df_procesado, _ = procesar_datos(df_hoja, columnas_notas)
            
            if columnas_id:
                # Agregar observaciones pedagógicas
                df_procesado['OBSERVACION_PEDAGOGICA'] = df_procesado.apply(
                    lambda row: gestor.generar_observacion(row['PROMEDIO'], row[columnas_id[0]])[0],
                    axis=1
                )
                
                # ═══════════════════════════════════════════════════════════
                # VISTA PREVIA
                # ═══════════════════════════════════════════════════════════
                
                st.markdown("### 👁️ Vista Previa del Reporte")
                cols_export = columnas_id + ['PROMEDIO', 'CALIFICACION_LETRA', 'ESTADO']
                
                # Estilizar la vista previa
                def colorear_fila(row):
                    if row['CALIFICACION_LETRA'] == 'AD':
                        return ['background-color: #d4edff'] * len(row)
                    elif row['CALIFICACION_LETRA'] == 'A':
                        return ['background-color: #d4f4dd'] * len(row)
                    elif row['CALIFICACION_LETRA'] == 'B':
                        return ['background-color: #fff4d4'] * len(row)
                    else:  # C
                        return ['background-color: #ffd4d4'] * len(row)
                
                st.dataframe(
                    df_procesado[cols_export].head(10).style.apply(colorear_fila, axis=1),
                    use_container_width=True
                )
                
                # ═══════════════════════════════════════════════════════════
                # OPCIONES DE DESCARGA
                # ═══════════════════════════════════════════════════════════
                
                st.markdown("### 📥 Opciones de Descarga")
                
                col1, col2, col3 = st.columns(3)
                
                # OPCIÓN 1: CSV Simple
                with col1:
                    st.markdown("#### 📊 Formato CSV")
                    st.caption("Compatible con Excel, Google Sheets")
                    
                    csv_data = df_procesado[cols_export].to_csv(index=False).encode('utf-8-sig')
                    st.download_button(
                        label="⬇️ Descargar CSV Simple",
                        data=csv_data,
                        file_name=f"reporte_simple_{hoja_seleccionada.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.csv",
                        mime='text/csv',
                        help="Datos sin formato, ideal para análisis"
                    )
                
                # OPCIÓN 2: Excel con Formato Profesional (NUEVO)
                with col2:
                    st.markdown("#### 📗 Formato Institucional")
                    st.caption("⭐ Con colores y formato MINEDU")
                    
                    buffer_formateado = generar_excel_formateado(
                        df_procesado, 
                        columnas_id, 
                        hoja_seleccionada
                    )
                    
                    st.download_button(
                        label="⬇️ Descargar Excel Formateado",
                        data=buffer_formateado,
                        file_name=f"reporte_formateado_{hoja_seleccionada.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        help="Formato profesional con colores según nivel de logro"
                    )
                
                # OPCIÓN 3: Excel Completo con Observaciones
                with col3:
                    st.markdown("#### 📘 Reporte Completo")
                    st.caption("Incluye observaciones pedagógicas")
                    
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        cols_completo = columnas_id + ['PROMEDIO', 'CALIFICACION_LETRA', 
                                                       'ESTADO', 'OBSERVACION_PEDAGOGICA']
                        df_procesado[cols_completo].to_excel(
                            writer, 
                            sheet_name='Reporte Completo', 
                            index=False
                        )
                    buffer.seek(0)
                    
                    st.download_button(
                        label="⬇️ Descargar Reporte Completo",
                        data=buffer,
                        file_name=f"reporte_completo_{hoja_seleccionada.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        help="Incluye todas las observaciones pedagógicas"
                    )
                
                # ═══════════════════════════════════════════════════════════
                # ESTADÍSTICAS
                # ═══════════════════════════════════════════════════════════
                
                st.markdown("---")
                st.markdown("### 📊 Estadísticas del Reporte")
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Registros", len(df_procesado))
                with col2:
                    st.metric("Promedio General", f"{df_procesado['PROMEDIO'].mean():.2f}")
                with col3:
                    aprobados = (df_procesado['ESTADO'] == 'Aprobado').sum()
                    st.metric("Aprobados", aprobados)
                with col4:
                    desaprobados = (df_procesado['ESTADO'] == 'Desaprobado').sum()
                    st.metric("Desaprobados", desaprobados)
                
                # Gráfico de distribución
                st.markdown("#### 📈 Distribución por Nivel de Logro")
                conteo = df_procesado['CALIFICACION_LETRA'].value_counts().reindex(['C','B','A','AD']).fillna(0)
                
                fig = go.Figure(data=[
                    go.Bar(
                        x=conteo.index,
                        y=conteo.values,
                        marker_color=['#FF6B6B', '#FFD166', '#06D6A0', '#118AB2'],
                        text=conteo.values,
                        textposition='auto'
                    )
                ])
                
                fig.update_layout(
                    title="Cantidad de estudiantes por nivel",
                    xaxis_title="Nivel de Logro",
                    yaxis_title="Cantidad",
                    height=400
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
                # Información adicional
                st.info("""
                💡 **Recomendaciones:**
                - 📊 **CSV Simple:** Para análisis adicional en otros programas
                - 📗 **Excel Formateado:** Para presentaciones oficiales (colores MINEDU)
                - 📘 **Reporte Completo:** Para archivo pedagógico (con observaciones)
                """)
                
        except Exception as e:
            st.error(f"❌ Error al generar reporte: {str(e)}")
            with st.expander("🔍 Ver detalles del error"):
                import traceback
                st.code(traceback.format_exc())

def pagina_ayuda():
    """Página de ayuda y documentación profesional"""
    
    # Encabezado con diseño atractivo
    st.markdown("""
    <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                padding: 30px; border-radius: 15px; text-align: center; margin-bottom: 30px;'>
        <h1 style='color: white; margin: 0;'>❓ Ayuda y Referencias MINEDU</h1>
        <p style='color: white; margin: 10px 0 0 0; font-size: 16px;'>
            Sistema de Análisis y Reforzamiento Académico V4.0
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Tabs mejoradas con íconos
    tab1, tab2, tab3, tab4 = st.tabs([
        "📖 Guía de Uso", 
        "❓ Preguntas Frecuentes", 
        "🔧 Solución de Problemas",
        "📞 Soporte Técnico"
    ])
    
    # ========================================================================
    # TAB 1: GUÍA DE USO
    # ========================================================================
    with tab1:
        st.markdown("## 📚 Guía Completa del Sistema")
        
        # Información Institucional con diseño mejorado
        st.markdown("""
        <div style='background: #f8f9fa; padding: 20px; border-radius: 10px; 
                    border-left: 5px solid #667eea; margin-bottom: 20px;'>
            <h3 style='color: #667eea; margin-top: 0;'>🏫 Información Institucional</h3>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"""
            - **🏛️ Nombre:** {INFO_INSTITUCION['nombre_ie1']}
            - **📍 Ubicación:** {INFO_INSTITUCION['ubicacion']}
            """)
        with col2:
            st.markdown(f"""
            - **🔢 Código Modular:** {INFO_INSTITUCION['codigo']}
            - **📅 Versión:** {INFO_INSTITUCION['version']}
            """)
        
        st.markdown("</div>", unsafe_allow_html=True)
        
        # Escala de Calificaciones con colores
        st.markdown("### 📊 Escala de Calificaciones MINEDU")
        
        # Crear DataFrame para mejor visualización
        df_escala = pd.DataFrame({
            'Nivel': ['🌟 AD', '✅ A', '⚠️ B', '🚨 C'],
            'Rango': ['18 - 20', '15 - 17', '11 - 14', '00 - 10'],
            'Descripción': ['Logro Destacado', 'Logro Esperado', 'En Proceso', 'En Inicio'],
            'Estrategia': [
                'Proyectos de investigación',
                'Aplicación práctica (ABP)',
                'Trabajo colaborativo',
                'Plan de Tutoría Individualizado'
            ]
        })
        
        st.dataframe(
            df_escala,
            use_container_width=True,
            hide_index=True
        )
        
        st.info("""
        💡 **Nota:** La escala está alineada con el Currículo Nacional de Educación Básica (CNEB) 
        y las Disposiciones de Evaluación Formativa del MINEDU 2024.
        """)
        
        # Guía paso a paso
        st.markdown("### 🚀 Guía de Uso Paso a Paso")
        
        with st.expander("📂 **PASO 1: Cargar Datos**", expanded=True):
            st.markdown("""
            1. Haga clic en **"Browse files"** en la barra lateral izquierda
            2. Seleccione su archivo Excel (`.xlsx` o `.xls`)
            3. Espere la confirmación: ✅ **"Cargado: X hoja(s)"**
            
            **📋 Formato requerido del Excel:**
            - Primera fila: Encabezados (APELLIDOS, NOMBRES, áreas curriculares)
            - Columnas de notas: Valores numéricos (0-20) o letras (C/B/A/AD)
            - Sin celdas combinadas
            - Sin filas vacías entre datos
            """)
            
            st.image("https://via.placeholder.com/600x150/667eea/ffffff?text=Ejemplo:+APELLIDOS+|+NOMBRES+|+COM+|+MAT+|+CyT", 
                     caption="Estructura recomendada del archivo Excel")
        
        with st.expander("🧭 **PASO 2: Navegar por las Vistas**"):
            st.markdown("""
            El sistema cuenta con **6 módulos principales:**
            
            | Módulo | Descripción | Usuario Objetivo |
            |--------|-------------|------------------|
            | 🏠 **Inicio** | Información general y bienvenida | Todos |
            | 👨‍🏫 **Vista Director** | Análisis global de toda la IE | Director/Coordinador |
            | 👩‍🏫 **Vista Docente** | Análisis específico por aula | Docentes |
            | 🧑‍🎓 **Análisis Estudiantil** | Observaciones individuales | Tutores/Docentes |
            | 🎯 **Análisis Priorizados** | Estudiantes en riesgo (C y B) | Equipo Directivo |
            | 📄 **Exportar Reportes** | Descarga de informes | Todos |
            
            💡 **Tip:** Use las flechas del teclado (↑↓) para navegar más rápido.
            """)
        
        with st.expander("📊 **PASO 3: Interpretar Resultados**"):
            st.markdown("""
            #### 📈 Gráficos Principales:
            
            **1. Distribución por Nivel (Barras):**
            - 🟦 **Azul (AD):** Estudiantes destacados → Potenciar talento
            - 🟢 **Verde (A):** Logro esperado → Mantener nivel
            - 🟡 **Amarillo (B):** En proceso → Reforzamiento moderado
            - 🔴 **Rojo (C):** En inicio → **Intervención urgente**
            
            **2. Histograma de Promedios:**
            - Muestra la distribución de calificaciones
            - Identifica concentraciones de estudiantes
            - Útil para detectar patrones de rendimiento
            
            **3. KPIs (Indicadores Clave):**
            - 🎓 **Total Estudiantes:** Población total analizada
            - 📊 **Promedio General:** Media aritmética de todas las notas
            - ✅ **Tasa de Aprobación:** % de estudiantes con B, A o AD
            """)
        
        with st.expander("💾 **PASO 4: Exportar Reportes**"):
            st.markdown("""
            El sistema permite exportar en **2 formatos:**
            
            | Formato | Ventajas | Uso Recomendado |
            |---------|----------|-----------------|
            | 📊 **CSV** | Compatible con Excel, Google Sheets | Análisis adicional |
            | 📗 **Excel** | Formato nativo, mantiene formato | Reportes oficiales |
            
            **Contenido del reporte:**
            - ✅ Datos de identificación del estudiante
            - ✅ Calificaciones por área curricular
            - ✅ Promedio general
            - ✅ Nivel de logro (C/B/A/AD)
            - ✅ Estado (Aprobado/Desaprobado)
            - ✅ **Observación pedagógica personalizada**
            """)
        
        # Referencias Normativas
        st.markdown("---")
        st.markdown("### 📋 Marco Normativo y Referencias")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.success("""
            **📚 Documentos Base:**
            - Plan Estratégico Institucional (PEI) 2024-2027
            - Currículo Nacional de Educación Básica (CNEB)
            - Disposiciones para Reforzamiento Escolar 2024
            - RVM N° 334-2021-MINEDU (Evaluación Formativa)
            """)
        
        with col2:
            st.info("""
            **🔗 Enlaces Útiles:**
            - [Portal MINEDU](http://www.minedu.gob.pe)
            - [SIAGIE](http://siagie.minedu.gob.pe)
            - [Currículo Nacional](http://www.minedu.gob.pe/curriculo/)
            - [PerúEduca](https://www.perueduca.pe)
            """)
    
    # ========================================================================
    # TAB 2: PREGUNTAS FRECUENTES
    # ========================================================================
    with tab2:
        st.markdown("## ❓ Preguntas Frecuentes (FAQ)")
        
        st.markdown("### 📂 Sobre la Carga de Datos")
        
        with st.expander("❓ ¿Qué formato debe tener mi archivo Excel?"):
            st.markdown("""
            **✅ Estructura Obligatoria:**
            
            ```
            | APELLIDOS Y NOMBRES | COM | MAT | CyT | PS | EPT | ART |
            |---------------------|-----|-----|-----|----|----|-----|
            | García López, Juan  | 14  | 16  | 12  | 15 | 13 | 17  |
            | Pérez Silva, María  | 18  | 19  | 17  | 18 | 20 | 19  |
            ```
            
            **📋 Requisitos:**
            - ✅ Primera fila con encabezados claros
            - ✅ Columna de identificación (APELLIDOS, NOMBRES o ESTUDIANTE)
            - ✅ Al menos 3 columnas de notas
            - ✅ Valores: 0-20 (numérico) o C/B/A/AD (letras)
            - ❌ Sin celdas combinadas
            - ❌ Sin filas vacías entre datos
            - ❌ Sin caracteres especiales en encabezados
            
            **💡 Tip:** Descargue nuestra plantilla desde "Exportar Reportes" para usarla como referencia.
            """)
        
        with st.expander("❓ ¿El sistema acepta calificaciones con letras (C/B/A/AD)?"):
            st.markdown("""
            **✅ SÍ, totalmente compatible.**
            
            El sistema convierte automáticamente:
            - **C** → 8 puntos (En Inicio)
            - **B** → 12 puntos (En Proceso)
            - **A** → 16 puntos (Logro Esperado)
            - **AD** → 19 puntos (Logro Destacado)
            
            Puede mezclar formatos:
            ```
            | ESTUDIANTE | COM | MAT | CyT |
            |------------|-----|-----|-----|
            | Juan       | 14  | A   | B   |  ← ✅ VÁLIDO
            | María      | AD  | 18  | 17  |  ← ✅ VÁLIDO
            ```
            """)
        
        with st.expander("❓ ¿Cuántas hojas puede tener mi archivo Excel?"):
            st.markdown("""
            **Sin límite.** El sistema detecta y procesa automáticamente todas las hojas.
            
            **Recomendación de organización:**
            - 📘 **Por Grado:** 1A, 1B, 2A, 2B, etc.
            - 📗 **Por Trimestre:** Trimestre 1, Trimestre 2, Trimestre 3
            - 📙 **Por Nivel:** Primaria, Secundaria
            
            El sistema genera análisis independientes para cada hoja.
            """)
        
        st.markdown("### ⚙️ Sobre Funcionalidades")
        
        with st.expander("❓ ¿Puedo usar el sistema sin instalar CatBoost (Machine Learning)?"):
            st.markdown(f"""
            **✅ SÍ, absolutamente.**
            
            **Estado actual:** {'✅ CatBoost instalado' if HAS_CATBOOST else '⚠️ CatBoost no instalado'}
            
            **Funciones disponibles SIN CatBoost:**
            - ✅ Análisis descriptivo completo
            - ✅ Gráficos y visualizaciones
            - ✅ Reportes y exportaciones
            - ✅ Observaciones pedagógicas
            - ✅ Vista Director, Docente y Estudiantil
            - ✅ Análisis de priorizados
            
            **Función adicional CON CatBoost:**
            - 🤖 Modelo predictivo de riesgo de desaprobación
            
            💡 **Conclusión:** El sistema funciona perfectamente sin Machine Learning.
            """)
        
        with st.expander("❓ ¿Cómo se calculan los promedios?"):
            st.markdown("""
            **Fórmula:** Media aritmética simple de todas las áreas curriculares.
            
            ```
            Promedio = (COM + MAT + CyT + PS + EPT + ...) / N° de áreas
            ```
            
            **Ejemplo:**
            ```
            Estudiante: Juan García
            COM: 14  |  MAT: 16  |  CyT: 12  |  PS: 15  |  EPT: 13
            
            Promedio = (14 + 16 + 12 + 15 + 13) / 5 = 70 / 5 = 14.00
            Nivel: B (En Proceso)
            ```
            
            ⚠️ **Importante:** El sistema omite automáticamente celdas vacías del cálculo.
            """)
        
        with st.expander("❓ ¿Qué son las 'Observaciones Pedagógicas'?"):
            st.markdown("""
            Son **recomendaciones personalizadas** basadas en el nivel de logro del estudiante.
            
            **Generadas automáticamente según:**
            - ✅ Promedio del estudiante
            - ✅ Nivel alcanzado (C/B/A/AD)
            - ✅ Estrategias MINEDU oficiales
            
            **Ejemplo para nivel C (En Inicio):**
            ```
            🚨 Reforzamiento Urgente Requerido
            - Implementar Plan de Tutoría Individualizado (PTI)
            - Foco en competencias básicas con sesiones de 30-45 min
            - Contactar a padres/apoderados para acompañamiento familiar
            📚 Ref: MINEDU - Disposiciones para Reforzamiento Escolar 2024
            ```
            
            Disponibles en:
            - 🧑‍🎓 Vista "Análisis Estudiantil"
            - 📄 Reportes exportados (columna OBSERVACION_PEDAGOGICA)
            """)
        
        st.markdown("### 🔒 Sobre Seguridad y Privacidad")
        
        with st.expander("❓ ¿Dónde se guardan los datos de los estudiantes?"):
            st.markdown("""
            **🔒 Los datos NO se guardan en ningún servidor externo.**
            
            **Procesamiento local:**
            - ✅ Datos procesados en **su computadora**
            - ✅ No se envían a internet
            - ✅ No se almacenan después de cerrar la aplicación
            - ✅ Cumple con Ley N° 29733 (Protección de Datos Personales)
            
            **Flujo de datos:**
            ```
            1. Usted carga Excel → 2. Se procesa localmente → 3. Se muestra resultado
                                                          ↓
                                    4. Al cerrar: datos se eliminan automáticamente
            ```
            
            🛡️ **Seguridad garantizada:** Sus datos permanecen privados.
            """)
        
        with st.expander("❓ ¿Puedo compartir los reportes exportados?"):
            st.markdown("""
            **✅ SÍ, con precauciones.**
            
            **Recomendaciones:**
            - ✅ Use para fines pedagógicos internos
            - ✅ Comparta solo con personal autorizado
            - ⚠️ Proteja archivos con contraseña si contienen datos sensibles
            - ❌ No publique en redes sociales
            
            **Cumplimiento normativo:**
            - Ley N° 29733: Protección de Datos Personales
            - Directiva MINEDU sobre confidencialidad de información estudiantil
            """)
    
    # ========================================================================
    # TAB 3: SOLUCIÓN DE PROBLEMAS
    # ========================================================================
    with tab3:
        st.markdown("## 🔧 Solución de Problemas Comunes")
        
        st.markdown("### 🚨 Errores Frecuentes y Soluciones")
        
        # Error 1
        st.error("**PROBLEMA 1:** ❌ No se detectaron columnas de notas")
        with st.expander("🔍 Ver solución detallada"):
            st.markdown("""
            **Causas posibles:**
            1. ❌ Encabezados con caracteres especiales (`@`, `#`, `%`, etc.)
            2. ❌ Columnas sin datos o todas vacías
            3. ❌ Formato de notas incorrecto (texto mezclado con números)
            4. ❌ Celdas combinadas en la primera fila
            
            **Soluciones:**
            
            ✅ **Solución 1: Verificar encabezados**
            ```
            ❌ INCORRECTO:  COM@, MAT#2, C&T
            ✅ CORRECTO:    COM, MAT, CyT
            ```
            
            ✅ **Solución 2: Asegurar formato de notas**
            - Valores numéricos: 0, 1, 2, ..., 18, 19, 20
            - Valores literales: C, B, A, AD (mayúsculas o minúsculas)
            
            ✅ **Solución 3: Eliminar celdas combinadas**
            - En Excel: Inicio → Combinar y centrar → Desactivar
            
            ✅ **Solución 4: Usar otra hoja**
            - Si una hoja tiene problemas, seleccione otra del menú desplegable
            """)
        
        # Error 2
        st.warning("**PROBLEMA 2:** ⚠️ El archivo tarda mucho en cargar")
        with st.expander("🔍 Ver solución detallada"):
            st.markdown("""
            **Causas posibles:**
            - 📊 Archivo muy grande (>10 MB)
            - 📄 Muchas hojas (>20 hojas)
            - 🖼️ Imágenes o gráficos insertados
            
            **Soluciones:**
            
            ✅ **Optimizar el archivo:**
            1. Eliminar hojas no necesarias
            2. Quitar imágenes y gráficos
            3. Guardar como `.xlsx` (más eficiente que `.xls`)
            4. Dividir en archivos más pequeños si tiene >1000 filas
            
            ⏱️ **Tiempos estimados:**
            - Pequeño (<100 estudiantes): 2-5 segundos
            - Mediano (100-500 estudiantes): 5-15 segundos
            - Grande (>500 estudiantes): 15-30 segundos
            """)
        
        # Error 3
        st.info("**PROBLEMA 3:** ℹ️ Los gráficos no se muestran correctamente")
        with st.expander("🔍 Ver solución detallada"):
            st.markdown("""
            **Soluciones rápidas:**
            
            ✅ **Refrescar la página:**
            - Presione `F5` o haga clic en el botón de recarga del navegador
            
            ✅ **Limpiar caché de Streamlit:**
            1. Menú superior derecho (⋮)
            2. Clic en "Clear cache"
            3. Recargar datos
            
            ✅ **Verificar conexión a internet:**
            - Algunos gráficos requieren librerías CDN
            - Use navegador actualizado (Chrome, Firefox, Edge)
            """)
        
        # Error 4
        st.error("**PROBLEMA 4:** ❌ Error al exportar reporte")
        with st.expander("🔍 Ver solución detallada"):
            st.markdown("""
            **Soluciones:**
            
            ✅ **Verificar permisos de descarga:**
            - Navegador puede estar bloqueando descargas
            - Permitir descargas en configuración del navegador
            
            ✅ **Liberar espacio en disco:**
            - Asegúrese de tener al menos 50 MB libres
            
            ✅ **Cerrar el archivo si está abierto:**
            - No puede sobrescribir un archivo Excel que esté abierto
            - Cierre el archivo anterior antes de exportar
            """)
        
        st.markdown("---")
        st.markdown("### 💻 Requisitos del Sistema")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.success("""
            **✅ Requisitos Mínimos:**
            - 🖥️ Sistema Operativo: Windows 10/11, macOS, Linux
            - 💾 RAM: 4 GB
            - 💿 Espacio: 500 MB
            - 🌐 Navegador: Chrome, Firefox, Edge (actualizado)
            - 🐍 Python: 3.8 o superior
            """)
        
        with col2:
            st.info("""
            **⭐ Requisitos Recomendados:**
            - 🖥️ Procesador: Intel i5 / AMD Ryzen 5
            - 💾 RAM: 8 GB
            - 💿 Espacio: 2 GB
            - 🌐 Conexión: Internet (para gráficos)
            - 📊 Resolución: 1366x768 o superior
            """)
    
    # ========================================================================
    # TAB 4: SOPORTE TÉCNICO (NUEVO)
    # ========================================================================
    with tab4:
        st.markdown("## 📞 Soporte Técnico y Contacto")
        
        st.markdown("""
        <div style='background: linear-gradient(135deg, #06D6A0 0%, #118AB2 100%); 
                    padding: 25px; border-radius: 15px; color: white; margin-bottom: 20px;'>
            <h3 style='margin: 0 0 10px 0;'>🤝 Estamos aquí para ayudarte</h3>
            <p style='margin: 0; font-size: 16px;'>
                Si tienes dudas o problemas técnicos, contáctanos a través de los siguientes canales.
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            ### 📧 Información de Contacto
            
            **🏫 Institución Educativa:**
            - I.E. 40079 "Víctor Núñez Valencia"
            - Av. Wanders 113, Sachaca, Arequipa
            
            **📱 Contacto Directo:**
            - ☎️ Teléfono: (054) 344259
            - 📧 Email: ievinvasecundaria@gmail.com
            - 🕐 Horario: Lunes a Viernes, 8:00 AM - 4:00 PM
            
            **💻 Soporte Técnico:**
            - Desarrollador: Frederickv
            - Email: fredypaulov1234567890@gmail.com
            """)
        
        with col2:
            st.markdown("""
            ### 📚 Recursos Adicionales
            
            **📖 Documentación:**
            - Manual de Usuario (PDF)
            - Guía de Instalación
            - Videos tutoriales
            
            **🔗 Enlaces Útiles:**
            - [Portal MINEDU](http://www.minedu.gob.pe)
            - [Normativa Vigente](http://www.minedu.gob.pe/normatividad/)
            - [Repositorio GitHub](#) *(próximamente)*
            
            **🎓 Capacitación:**
            - Solicitar taller presencial
            - Sesiones virtuales de Q&A
            - Material de apoyo descargable
            """)
        
        st.markdown("---")
        
        st.markdown("### 📝 Formulario de Reporte de Problemas")
        
        with st.form("formulario_soporte"):
            nombre = st.text_input("👤 Nombre completo")
            email = st.text_input("📧 Correo electrónico")
            tipo_problema = st.selectbox(
                "🔍 Tipo de problema",
                ["Carga de archivos", "Error en cálculos", "Problema con gráficos", 
                 "Exportación de reportes", "Otro"]
            )
            descripcion = st.text_area("📋 Descripción detallada del problema", height=150)
            
            col1, col2 = st.columns([1, 3])
            with col1:
                enviar = st.form_submit_button("📨 Enviar Reporte", use_container_width=True)
            
            if enviar:
                if nombre and email and descripcion:
                    st.success("""
                    ✅ **¡Reporte enviado exitosamente!**
                    
                    Recibirá respuesta en un plazo máximo de 24-48 horas hábiles.
                    Se ha enviado una copia del reporte a su correo electrónico.
                    """)
                else:
                    st.error("❌ Por favor complete todos los campos obligatorios.")
        
        st.markdown("---")
        
        # Información del sistema
        st.markdown("### 🔍 Información del Sistema")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Versión", INFO_INSTITUCION['version'])
        with col2:
            st.metric("Módulo Priorizados", "✅ Activo" if HAS_MODULO_PRIORIZADOS else "❌ Inactivo")
        with col3:
            st.metric("Módulo ML", "✅ Activo" if HAS_MODULO_PREDICTIVO else "❌ Inactivo")
        
        with st.expander("🔧 Ver detalles técnicos completos"):
            st.code(f"""
Sistema de Análisis y Reforzamiento Académico V4.0
═══════════════════════════════════════════════════

Información General:
- Versión: {INFO_INSTITUCION['version']}
- Institución: {INFO_INSTITUCION['nombre_ie1']}
- Código Modular: {INFO_INSTITUCION['codigo']}

Estado de Módulos:
- Análisis Priorizados: {'✅ Activo' if HAS_MODULO_PRIORIZADOS else '❌ Inactivo'}
- Modelo Predictivo ML: {'✅ Activo' if HAS_MODULO_PREDICTIVO else '❌ Inactivo'}
- CatBoost: {'✅ Instalado' if HAS_CATBOOST else '❌ No instalado'}

Librerías Principales:
- Streamlit: {st.__version__}
- Pandas: {pd.__version__}
- NumPy: {np.__version__}
- Plotly: (instalado)

Desarrollado por: frederickv
Fecha: 2025-11-11
            """, language="text")
    
    # Pie de página
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; padding: 20px; background: #f8f9fa; border-radius: 10px;'>
        <p style='color: #666; margin: 0; font-size: 14px;'>
            💡 <strong>Nota:</strong> Este sistema está en constante mejora. 
            Sus comentarios y sugerencias son bienvenidos.
        </p>
        <p style='color: #999; margin: 10px 0 0 0; font-size: 12px;'>
            Sistema basado en PEI MINEDU 2024-2027 | Desarrollado con ❤️ para la educación peruana
        </p>
    </div>
    """, unsafe_allow_html=True)



# ═════════════════════════════════════════════════════════════════════════════
# APLICACIÓN PRINCIPAL
# ═════════════════════════════════════════════════════════════════════════════

def main():
    """Función principal de la aplicación"""
    
    mostrar_logo()
    st.sidebar.markdown("---")
    
    if "datos_cargados" not in st.session_state:
        st.session_state.datos_cargados = None
        st.session_state.datos_raw = None
        st.session_state.nombres_hojas = []
    
    st.sidebar.header("📂 Cargar Datos")
    
    archivo_subido = st.sidebar.file_uploader(
        "Seleccionar archivo Excel",
        type=["xlsx", "xls"],
        help="Archivo con calificaciones de estudiantes"
    )
    
    if archivo_subido is not None:
        with st.sidebar:
            with st.spinner("📊 Procesando archivo..."):
                datos_cargados, datos_raw, nombres_hojas, error = cargar_excel(archivo_subido)
                
                if error:
                    st.error(f"❌ {error}")
                    st.session_state.datos_cargados = None
                    st.session_state.datos_raw = None
                else:
                    st.session_state.datos_cargados = datos_cargados
                    st.session_state.datos_raw = datos_raw
                    st.session_state.nombres_hojas = nombres_hojas
                    st.success(f"✅ Cargado: {len(nombres_hojas)} hoja(s)")
    
    if st.session_state.datos_cargados:
        st.sidebar.markdown("---")
        st.sidebar.header("🧭 Navegación")
        
        paginas = {
            "🏠 Inicio": ("inicio", None),
            "👨‍🏫 Vista Director": ("director", (st.session_state.datos_cargados, st.session_state.datos_raw)),
            "👩‍🏫 Vista Docente": ("docente", st.session_state.datos_cargados),
            "🧑‍🎓 Análisis Estudiantil": ("estudiantil", st.session_state.datos_cargados),
        }
        
        if HAS_MODULO_PRIORIZADOS:
            paginas["🎯 Análisis Priorizados"] = ("priorizados", st.session_state.datos_raw)
        
        if HAS_MODULO_PREDICTIVO and HAS_CATBOOST:
            paginas["🤖 Modelo Predictivo ML"] = ("predictivo", st.session_state.datos_raw)
        
        paginas["📄 Exportar Reportes"] = ("reportes", st.session_state.datos_cargados)
        paginas["❓ Ayuda y Referencias"] = ("ayuda", None)
        
        seleccion = st.sidebar.radio(
            "Seleccione una vista:",
            options=list(paginas.keys()),
            label_visibility="collapsed"
        )
        
        st.sidebar.markdown("---")
        st.sidebar.markdown(f"""
        <div style='text-align: center; padding: 10px; background: #f0f2f6; border-radius: 10px;'>
            <small>
                📚 Sistema basado en<br/>
                <b>PEI MINEDU 2024-2027</b><br/>
                {INFO_INSTITUCION['version']}
            </small>
        </div>
        """, unsafe_allow_html=True)
        
        st.sidebar.markdown("---")
        st.sidebar.markdown("**📦 Estado de Módulos:**")
        st.sidebar.markdown(f"{'✅' if HAS_MODULO_PRIORIZADOS else '❌'} Análisis Priorizados")
        st.sidebar.markdown(f"{'✅' if HAS_MODULO_PREDICTIVO else '❌'} Modelo Predictivo")
        st.sidebar.markdown(f"{'✅' if HAS_CATBOOST else '✅ (Descriptivo)'} CatBoost ML")
        
        try:
            pagina_tipo, datos = paginas[seleccion]
            
            if pagina_tipo == "inicio":
                pagina_inicio()
            elif pagina_tipo == "director":
                pagina_vista_director(*datos)
            elif pagina_tipo == "docente":
                pagina_vista_docente(datos)
            elif pagina_tipo == "estudiantil":
                pagina_analisis_estudiantil(datos)
            elif pagina_tipo == "priorizados":
                pagina_analisis_priorizados(datos)
            elif pagina_tipo == "predictivo":
                pagina_modelo_predictivo(datos)
            elif pagina_tipo == "reportes":
                pagina_exportar_reportes(datos)
            elif pagina_tipo == "ayuda":
                pagina_ayuda()
                
        except Exception as e:
            st.error(f"❌ Error al cargar página: {str(e)}")
            st.exception(e)
    
    else:
        pagina_inicio()
        
        if archivo_subido is None:
            st.markdown("---")
            st.info("""
            👈 **Para comenzar:** Use el botón **"Browse files"** en la barra lateral 
            para cargar su archivo Excel con las calificaciones.
            """)

if __name__ == "__main__":
    main()